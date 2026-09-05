from typing import Optional, Union
import base64
import difflib
import io
import re
from datetime import datetime, date as date_type, timezone
from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session as DBSession
from openpyxl import load_workbook

from backend.database import get_db
from backend import models
from backend.services import openai_service
from backend.services.event_normalizer import (
    canonicalize_event, event_parts, sync_meet_session_events, upsert_meet_entry,
)
from backend.services.importer import parse_time_to_seconds

router = APIRouter()

LEVELS = ["club", "regional", "national", "international"]


class MeetCreate(BaseModel):
    name: str
    start_date: Optional[str] = Field(default=None, alias="date")
    end_date: Optional[str] = Field(default=None, alias="date_to")
    location: Optional[str] = None
    course: Optional[str] = None       # SCM / LCM
    level: Optional[str] = None        # club / regional / national / international
    warm_up_time: Optional[str] = None # "07:30"
    notes: Optional[str] = None

    class Config:
        populate_by_name = True


class MeetTargetCreate(BaseModel):
    swimmer_id: int
    events: list[str] = []
    priority: Optional[str] = None     # A / B / C
    target_times: Optional[dict] = None  # {"100 Freestyle": "52.50"}
    notes: Optional[str] = None


class MeetTargetUpdate(BaseModel):
    events: Optional[list[str]] = None
    priority: Optional[str] = None
    target_times: Optional[dict] = None
    notes: Optional[str] = None


class MeetSessionIn(BaseModel):
    name: str
    date: Optional[date_type] = None
    warm_up_time: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    order_index: int = 0
    events: list = Field(default_factory=list)
    notes: Optional[str] = None


class MeetImportRow(BaseModel):
    row_number: Optional[int] = None
    name: str
    date: date_type
    date_to: Optional[date_type] = None
    location: Optional[str] = None
    course: Optional[str] = None
    level: Optional[str] = None
    warm_up_time: Optional[str] = None
    notes: Optional[str] = None
    include: bool = True


class MeetImportConfirm(BaseModel):
    rows: list[MeetImportRow] = Field(default_factory=list)


_MEET_IMPORT_HEADERS = {
    "date": {"date start", "start date", "date from", "from"},
    "date_to": {"date end", "end date", "date to", "to"},
    "name": {"competition", "competition name", "gala", "gala name", "meet", "meet name"},
    "location": {"venue", "location", "pool"},
}


def _normalise_import_header(value) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())).strip()


def _clean_import_text(value) -> Optional[str]:
    if value is None:
        return None
    cleaned = re.sub(r"\s+", " ", str(value)).strip()
    return cleaned or None


def _parse_import_date(value) -> Optional[date_type]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_type):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _existing_meet_key(name: str, start_date: date_type) -> tuple[str, date_type]:
    return (re.sub(r"\s+", " ", name).strip().casefold(), start_date)


def _extract_meet_workbook(content: bytes, db: DBSession) -> dict:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="The workbook could not be opened as an .xlsx file") from exc

    worksheet = workbook.active
    header_row = None
    columns = {}
    for row_number in range(1, min(worksheet.max_row, 20) + 1):
        found = {}
        for column_number in range(1, worksheet.max_column + 1):
            header = _normalise_import_header(worksheet.cell(row_number, column_number).value)
            for field, aliases in _MEET_IMPORT_HEADERS.items():
                if header in aliases:
                    found[field] = column_number
                    break
        if {"date", "name"}.issubset(found):
            header_row = row_number
            columns = found
            break

    if header_row is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find the required Competition and Date Start columns",
        )

    existing = {
        _existing_meet_key(meet.name, meet.date): meet.id
        for meet in db.query(models.Meet).filter(models.Meet.date.is_not(None)).all()
    }
    upload_keys = set()
    rows = []
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        raw_name = worksheet.cell(row_number, columns["name"]).value
        raw_start = worksheet.cell(row_number, columns["date"]).value
        raw_end = worksheet.cell(row_number, columns["date_to"]).value if columns.get("date_to") else None
        raw_location = worksheet.cell(row_number, columns["location"]).value if columns.get("location") else None

        if all(value in (None, "") for value in (raw_name, raw_start, raw_end, raw_location)):
            continue

        name = _clean_import_text(raw_name)
        start_date = _parse_import_date(raw_start)
        end_date = _parse_import_date(raw_end) if raw_end not in (None, "") else start_date
        location = _clean_import_text(raw_location)
        course = "SCM" if name and re.search(r"\(\s*25m\s*\)", name, re.IGNORECASE) else (
            "LCM" if name and re.search(r"\(\s*50m\s*\)", name, re.IGNORECASE) else None
        )
        errors = []
        warnings = []
        if not name:
            errors.append("Competition name is missing")
        if not start_date:
            errors.append("Start date is missing or invalid")
        if raw_end not in (None, "") and not end_date:
            errors.append("End date is invalid")
        if start_date and end_date and end_date < start_date:
            errors.append("End date is before the start date")

        duplicate_id = None
        if name and start_date:
            key = _existing_meet_key(name, start_date)
            duplicate_id = existing.get(key)
            if duplicate_id:
                warnings.append("Already exists and will be skipped")
            elif key in upload_keys:
                warnings.append("Repeated in this workbook and will be skipped")
            else:
                upload_keys.add(key)

        can_import = not errors and not warnings
        rows.append({
            "row_number": row_number,
            "name": name or "",
            "date": start_date.isoformat() if start_date else None,
            "date_to": end_date.isoformat() if end_date else None,
            "location": location,
            "course": course,
            "level": None,
            "warm_up_time": None,
            "notes": None,
            "can_import": can_import,
            "existing_meet_id": duplicate_id,
            "errors": errors,
            "warnings": warnings,
        })

    if not rows:
        raise HTTPException(status_code=400, detail="No competition rows were found below the headings")

    return {
        "filename": None,
        "sheet_name": worksheet.title,
        "rows": rows,
        "summary": {
            "total": len(rows),
            "ready": sum(1 for row in rows if row["can_import"]),
            "duplicates": sum(1 for row in rows if row["warnings"]),
            "invalid": sum(1 for row in rows if row["errors"]),
        },
    }


@router.get("")
def list_meets(db: DBSession = Depends(get_db)):
    meets = db.query(models.Meet).order_by(models.Meet.date).all()
    return [_meet_summary(m, db) for m in meets]


@router.post("", status_code=201)
def create_meet(body: MeetCreate, db: DBSession = Depends(get_db)):
    data = body.model_dump(by_alias=True)
    # Parse date strings to date objects (by_alias already mapped start_date -> date, end_date -> date_to)
    if data.get('date'):
        data['date'] = datetime.fromisoformat(data['date']).date()
    if data.get('date_to'):
        data['date_to'] = datetime.fromisoformat(data['date_to']).date()
    meet = models.Meet(**data)
    db.add(meet)
    db.commit()
    db.refresh(meet)
    return _meet_summary(meet, db)


@router.post("/import/excel")
async def preview_meet_excel(file: UploadFile = File(...), db: DBSession = Depends(get_db)):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx workbook")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Workbook is too large (maximum 5 MB)")
    preview = _extract_meet_workbook(content, db)
    preview["filename"] = file.filename
    return preview


@router.post("/import/excel/confirm")
def confirm_meet_excel(body: MeetImportConfirm, db: DBSession = Depends(get_db)):
    created = []
    skipped = []
    seen = set()
    for row in body.rows:
        if not row.include:
            continue
        name = _clean_import_text(row.name)
        if not name:
            skipped.append({"row_number": row.row_number, "reason": "Competition name is missing"})
            continue
        end_date = row.date_to or row.date
        if end_date < row.date:
            skipped.append({"row_number": row.row_number, "reason": "End date is before the start date"})
            continue
        key = _existing_meet_key(name, row.date)
        same_day = db.query(models.Meet).filter(models.Meet.date == row.date).all()
        duplicate = next((meet for meet in same_day if _existing_meet_key(meet.name, meet.date) == key), None)
        if duplicate or key in seen:
            skipped.append({
                "row_number": row.row_number,
                "reason": "Already exists" if duplicate else "Repeated in import",
                "existing_meet_id": duplicate.id if duplicate else None,
            })
            continue
        seen.add(key)
        course = row.course.upper() if row.course and row.course.upper() in {"SCM", "LCM"} else None
        level = row.level.lower() if row.level and row.level.lower() in LEVELS else None
        meet = models.Meet(
            name=name,
            date=row.date,
            date_to=end_date,
            location=_clean_import_text(row.location),
            course=course,
            level=level,
            warm_up_time=_clean_import_text(row.warm_up_time),
            notes=_clean_import_text(row.notes),
        )
        db.add(meet)
        db.flush()
        created.append({"id": meet.id, "name": meet.name, "date": meet.date.isoformat()})
    db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "created_count": len(created),
        "skipped_count": len(skipped),
    }


@router.get("/{meet_id}")
def get_meet(meet_id: int, db: DBSession = Depends(get_db)):
    meet = _get_or_404(meet_id, db)
    return _meet_detail(meet, db)


@router.put("/{meet_id}")
def update_meet(meet_id: int, body: dict = Body(...), db: DBSession = Depends(get_db)):
    meet = _get_or_404(meet_id, db)
    allowed = {"name", "date", "date_to", "location", "course", "level", "warm_up_time", "notes"}
    for k, v in body.items():
        if k not in allowed:
            continue
        # Dates arrive as ISO strings from the client and must be real dates.
        setattr(meet, k, _parse_import_date(v) if k in {"date", "date_to"} else v)
    db.commit()
    return _meet_detail(meet, db)


@router.delete("/{meet_id}", status_code=204)
def delete_meet(meet_id: int, db: DBSession = Depends(get_db)):
    meet = _get_or_404(meet_id, db)
    db.delete(meet)
    db.commit()


# ---------------------------------------------------------------------------
# Targets (swimmer assignments)
# ---------------------------------------------------------------------------

@router.post("/{meet_id}/targets", status_code=201)
def add_target(meet_id: int, body: MeetTargetCreate, db: DBSession = Depends(get_db)):
    _get_or_404(meet_id, db)
    # Upsert — if swimmer already assigned, update instead
    existing = db.query(models.MeetTarget).filter(
        models.MeetTarget.meet_id == meet_id,
        models.MeetTarget.swimmer_id == body.swimmer_id,
    ).first()
    if existing:
        if body.events is not None:
            existing.events = body.events
        if body.priority is not None:
            existing.priority = body.priority
        if body.target_times is not None:
            existing.target_times = body.target_times
        if body.notes is not None:
            existing.notes = body.notes
        _sync_target_entries(existing, db)
        db.commit()
        db.refresh(existing)
        return _target_out(existing, db)

    target = models.MeetTarget(meet_id=meet_id, **body.model_dump())
    db.add(target)
    db.flush()
    _sync_target_entries(target, db)
    db.commit()
    db.refresh(target)
    return _target_out(target, db)


@router.put("/{meet_id}/targets/{target_id}")
def update_target(meet_id: int, target_id: int, body: MeetTargetUpdate, db: DBSession = Depends(get_db)):
    target = db.query(models.MeetTarget).filter(
        models.MeetTarget.id == target_id,
        models.MeetTarget.meet_id == meet_id,
    ).first()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    if body.events is not None:
        target.events = body.events
    if body.priority is not None:
        target.priority = body.priority
    if body.target_times is not None:
        target.target_times = body.target_times
    if body.notes is not None:
        target.notes = body.notes
    _sync_target_entries(target, db)
    db.commit()
    return _target_out(target, db)


@router.delete("/{meet_id}/targets/{target_id}", status_code=204)
def remove_target(meet_id: int, target_id: int, db: DBSession = Depends(get_db)):
    target = db.query(models.MeetTarget).filter(
        models.MeetTarget.id == target_id,
        models.MeetTarget.meet_id == meet_id,
    ).first()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    db.query(models.MeetEntry).filter(
        models.MeetEntry.meet_id == meet_id,
        models.MeetEntry.swimmer_id == target.swimmer_id,
    ).delete()
    db.delete(target)
    db.commit()


# ---------------------------------------------------------------------------
# Persisted competition timetable
# ---------------------------------------------------------------------------

@router.get("/{meet_id}/timetable")
def get_timetable(meet_id: int, db: DBSession = Depends(get_db)):
    _get_or_404(meet_id, db)
    rows = db.query(models.MeetSession).filter(
        models.MeetSession.meet_id == meet_id,
    ).order_by(models.MeetSession.date, models.MeetSession.order_index, models.MeetSession.start_time).all()
    return [_session_out(row) for row in rows]


@router.post("/{meet_id}/timetable", status_code=201)
def create_timetable_session(meet_id: int, body: MeetSessionIn, db: DBSession = Depends(get_db)):
    _get_or_404(meet_id, db)
    row = models.MeetSession(meet_id=meet_id, **body.model_dump())
    db.add(row)
    db.flush()
    sync_meet_session_events(row, db)
    db.commit()
    db.refresh(row)
    return _session_out(row)


@router.put("/{meet_id}/timetable/{session_id}")
def update_timetable_session(meet_id: int, session_id: int, body: MeetSessionIn, db: DBSession = Depends(get_db)):
    row = db.query(models.MeetSession).filter(
        models.MeetSession.id == session_id,
        models.MeetSession.meet_id == meet_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Timetable session not found")
    for key, value in body.model_dump().items():
        setattr(row, key, value)
    db.flush()
    sync_meet_session_events(row, db)
    db.commit()
    db.refresh(row)
    return _session_out(row)


@router.delete("/{meet_id}/timetable/{session_id}", status_code=204)
def delete_timetable_session(meet_id: int, session_id: int, db: DBSession = Depends(get_db)):
    row = db.query(models.MeetSession).filter(
        models.MeetSession.id == session_id,
        models.MeetSession.meet_id == meet_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Timetable session not found")
    event_ids = [event.id for event in row.normalized_events]
    if event_ids:
        db.query(models.MeetEntry).filter(models.MeetEntry.meet_event_id.in_(event_ids)).update(
            {"meet_event_id": None}, synchronize_session=False,
        )
    db.delete(row)
    db.commit()


@router.post("/{meet_id}/timetable/import")
def import_timetable(meet_id: int, body: dict = Body(...), db: DBSession = Depends(get_db)):
    """Replace or append timetable sessions from reviewed extraction output."""
    meet = _get_or_404(meet_id, db)
    sessions = body.get("sessions") or _sessions_from_extraction(body, meet)
    if not sessions:
        raise HTTPException(status_code=422, detail="No timetable sessions supplied")
    if body.get("replace", True):
        old_sessions = db.query(models.MeetSession).filter(models.MeetSession.meet_id == meet_id).all()
        old_event_ids = [event.id for session in old_sessions for event in session.normalized_events]
        if old_event_ids:
            db.query(models.MeetEntry).filter(models.MeetEntry.meet_event_id.in_(old_event_ids)).update(
                {"meet_event_id": None}, synchronize_session=False,
            )
        for old_session in old_sessions:
            db.delete(old_session)
        db.flush()
    created = []
    for index, item in enumerate(sessions):
        raw_date = item.get("date")
        parsed_date = None
        if raw_date and raw_date != "TBD":
            try:
                parsed_date = datetime.fromisoformat(str(raw_date)).date()
            except ValueError:
                parsed_date = None
        row = models.MeetSession(
            meet_id=meet_id,
            name=item.get("name") or (f"Session {index + 1}"),
            date=parsed_date,
            warm_up_time=item.get("warm_up_time"),
            start_time=item.get("start_time"),
            end_time=item.get("end_time"),
            order_index=item.get("order_index", index),
            events=_normalise_events(item.get("events") or []),
            notes=item.get("notes"),
        )
        db.add(row)
        db.flush()
        sync_meet_session_events(row, db)
        created.append(row)
    db.commit()
    return [_session_out(row) for row in created]


# ---------------------------------------------------------------------------
# Document extraction
# ---------------------------------------------------------------------------

@router.post("/{meet_id}/extract-schedule")
async def extract_schedule(
    meet_id: int,
    schedule_file: UploadFile = File(...),
    db: DBSession = Depends(get_db)
):
    """
    Extract events and dates from a meet schedule document.
    """
    meet = _get_or_404(meet_id, db)

    schedule_content = await schedule_file.read()
    mime_type = schedule_file.content_type or "application/pdf"

    # Build date context for the prompt
    date_context = ""
    if meet.date:
        date_context = f"The gala runs from {meet.date}"
        if meet.date_to and meet.date_to != meet.date:
            date_context += f" to {meet.date_to}"
        date_context += "."

    if mime_type.startswith("image/"):
        schedule_data = openai_service.parse_schedule_image(schedule_content, mime_type, date_context)
    else:
        schedule_data = openai_service.parse_schedule_document(schedule_content, date_context)

    return {
        "events": schedule_data.get("events", []),
        "by_date": schedule_data.get("by_date", {}),
        "sessions": schedule_data.get("sessions", []),
        "meet_id": meet_id,
    }


@router.post("/{meet_id}/extract-entries")
async def extract_entries(
    meet_id: int,
    entries_file: UploadFile = File(...),
    db: DBSession = Depends(get_db)
):
    """
    Extract swimmer names and their registered events from an entries document.
    """
    meet = _get_or_404(meet_id, db)

    # Get all squad swimmers for cross-referencing
    swimmers = db.query(models.Swimmer).filter(models.Swimmer.status == 'active').all()
    swimmer_names = {s.name.lower(): s for s in swimmers}

    entries_content = await entries_file.read()
    mime_type = entries_file.content_type or "text/plain"

    if mime_type.startswith("image/"):
        entries_data = openai_service.parse_entries_image(entries_content, mime_type)
    else:
        text_content = entries_content.decode('utf-8', errors='ignore')
        entries_data = openai_service.parse_entries_text(text_content)

    # Cross-reference swimmers with fuzzy matching
    extracted_swimmers = {}
    unmatched = []

    for swimmer_entry in entries_data.get("swimmers", []):
        name = swimmer_entry.get("name", "").strip()
        events = swimmer_entry.get("events", [])

        if not name:
            continue

        name_lower = name.lower()
        best_match = None
        best_score = 0.6

        for squad_name, swimmer in swimmer_names.items():
            score = difflib.SequenceMatcher(None, name_lower, squad_name).ratio()
            if score > best_score:
                best_score = score
                best_match = swimmer.name

        if best_match:
            if best_match not in extracted_swimmers:
                extracted_swimmers[best_match] = []
            extracted_swimmers[best_match].extend(events)
        else:
            unmatched.append({"name": name, "events": events, "match_score": best_score})

    return {
        "swimmers": extracted_swimmers,
        "unmatched": unmatched,
        "meet_id": meet_id,
    }


@router.post("/{meet_id}/combine-extractions")
async def combine_extractions(
    meet_id: int,
    body: dict = Body(...),
    db: DBSession = Depends(get_db)
):
    """
    Combine schedule and entries extractions and optionally auto-assign swimmers.
    Expects: {
        "events": [...],
        "swimmers": {"name": [...events]},
        "auto_assign": bool
    }
    """
    meet = _get_or_404(meet_id, db)

    events = body.get("events", [])
    swimmers_to_assign = body.get("swimmers", {})
    auto_assign = body.get("auto_assign", False)

    result = {
        "events": events,
        "swimmers": swimmers_to_assign,
        "assigned": []
    }

    if auto_assign:
        swimmers = db.query(models.Swimmer).all()
        swimmer_map = {s.name: s for s in swimmers}

        for swimmer_name, events_list in swimmers_to_assign.items():
            if swimmer_name in swimmer_map:
                try:
                    # Create or update meet target
                    existing = db.query(models.MeetTarget).filter(
                        models.MeetTarget.meet_id == meet_id,
                        models.MeetTarget.swimmer_id == swimmer_map[swimmer_name].id,
                    ).first()

                    if existing:
                        existing.events = list(set(existing.events or []) | set(events_list))
                        _sync_target_entries(existing, db, source="document")
                    else:
                        target = models.MeetTarget(
                            meet_id=meet_id,
                            swimmer_id=swimmer_map[swimmer_name].id,
                            events=events_list,
                            priority="B"
                        )
                        db.add(target)
                        db.flush()
                        _sync_target_entries(target, db, source="document")

                    result["assigned"].append(swimmer_name)
                except Exception as e:
                    pass

        db.commit()

    return result


# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------

class MeetResultRow(BaseModel):
    swimmer_id: int
    event: str
    time: Optional[str] = None          # "1:02.45" / "62.45" — blank clears the result
    round: Optional[str] = None         # Heat / Final / Semi
    date: Optional[str] = None          # defaults to the meet start date


class MeetResultsSubmit(BaseModel):
    results: list[MeetResultRow]


@router.get("/{meet_id}/results")
def get_meet_results(meet_id: int, db: DBSession = Depends(get_db)):
    """Every swim expected at this meet, pre-filled with any result already recorded."""
    meet = _get_or_404(meet_id, db)
    recorded = db.query(models.SwimTime).filter(models.SwimTime.meet_id == meet.id).all()
    recorded_by_key = {
        (row.swimmer_id, canonicalize_event(row.event), row.round or ""): row
        for row in recorded
    }

    swimmer_ids = {row.swimmer_id for row in recorded}
    entries = db.query(models.MeetEntry).filter(models.MeetEntry.meet_id == meet.id).all()
    targets = db.query(models.MeetTarget).filter(models.MeetTarget.meet_id == meet.id).all()
    swimmer_ids |= {entry.swimmer_id for entry in entries}
    swimmer_ids |= {target.swimmer_id for target in targets}
    swimmers = {
        row.id: row for row in
        db.query(models.Swimmer).filter(models.Swimmer.id.in_(swimmer_ids)).all()
    } if swimmer_ids else {}

    # Expected swims come from explicit entries first, then any target events
    # that were never turned into entries.
    expected: dict[tuple[int, str], dict] = {}
    for entry in entries:
        key = (entry.swimmer_id, entry.canonical_event)
        expected.setdefault(key, {
            "swimmer_id": entry.swimmer_id,
            "event": entry.event_name,
            "canonical_event": entry.canonical_event,
            "entry_time": entry.entry_time,
            "target_time": entry.target_time,
        })
    for target in targets:
        for event in target.events or []:
            key = (target.swimmer_id, canonicalize_event(event))
            expected.setdefault(key, {
                "swimmer_id": target.swimmer_id,
                "event": event,
                "canonical_event": canonicalize_event(event),
                "entry_time": None,
                "target_time": (target.target_times or {}).get(event),
            })

    rows = []
    for (swimmer_id, canonical), row in expected.items():
        swimmer = swimmers.get(swimmer_id)
        if not swimmer:
            continue
        result = recorded_by_key.pop((swimmer_id, canonical, ""), None)
        rows.append({
            **row,
            "swimmer_name": swimmer.name,
            "squad": swimmer.squad,
            "round": result.round if result else None,
            "recorded_time": _time_display(result.time_seconds) if result else None,
            "recorded_time_id": result.id if result else None,
            "unexpected": False,
        })

    # Results recorded against swims that were never entered — keep them visible
    # so a correction is possible rather than silently orphaned.
    for (swimmer_id, canonical, round_name), result in recorded_by_key.items():
        swimmer = swimmers.get(swimmer_id)
        if not swimmer:
            continue
        rows.append({
            "swimmer_id": swimmer_id,
            "swimmer_name": swimmer.name,
            "squad": swimmer.squad,
            "event": result.event,
            "canonical_event": canonical,
            "entry_time": None,
            "target_time": None,
            "round": result.round,
            "recorded_time": _time_display(result.time_seconds),
            "recorded_time_id": result.id,
            "unexpected": True,
        })

    rows.sort(key=lambda row: (row["swimmer_name"] or "", row["event"] or ""))
    return {
        "meet": _meet_summary(meet, db),
        "rows": rows,
        "recorded_count": sum(1 for row in rows if row["recorded_time"]),
    }


@router.post("/{meet_id}/results")
def save_meet_results(meet_id: int, body: MeetResultsSubmit, db: DBSession = Depends(get_db)):
    """Record race times against this meet and into each swimmer's times history."""
    meet = _get_or_404(meet_id, db)
    default_date = meet.date
    saved, cleared, errors = 0, 0, []

    for row in body.results:
        swimmer = db.query(models.Swimmer).filter(models.Swimmer.id == row.swimmer_id).first()
        if not swimmer:
            errors.append(f"Swimmer {row.swimmer_id} not found")
            continue

        canonical = canonicalize_event(row.event)
        existing = next(
            (item for item in db.query(models.SwimTime).filter(
                models.SwimTime.meet_id == meet.id,
                models.SwimTime.swimmer_id == row.swimmer_id,
            ).all()
             if canonicalize_event(item.event) == canonical
             and (item.round or "") == (row.round or "")),
            None,
        )

        raw_time = (row.time or "").strip()
        if not raw_time:
            if existing:
                db.delete(existing)
                cleared += 1
            continue

        seconds = parse_time_to_seconds(raw_time)
        if seconds is None or seconds <= 0:
            errors.append(f"{swimmer.name} - {row.event}: could not read the time '{raw_time}'")
            continue

        swim_date = _parse_import_date(row.date) or default_date
        distance, stroke = event_parts(row.event)
        event_base = canonical.title()
        event_label = f"{event_base} {meet.course}" if meet.course else event_base

        if existing:
            existing.time_seconds = seconds
            existing.date = swim_date
            existing.event = event_label
            existing.stroke = stroke.title() if stroke else None
            existing.distance = distance
            existing.course = meet.course
            existing.venue = meet.location
            existing.level = meet.level
            existing.meet = meet.name
        else:
            db.add(models.SwimTime(
                swimmer_id=row.swimmer_id,
                event=event_label,
                stroke=stroke.title() if stroke else None,
                course=meet.course,
                distance=distance,
                time_seconds=seconds,
                date=swim_date,
                meet=meet.name,
                venue=meet.location,
                level=meet.level,
                round=row.round or None,
                source="meet_results",
                meet_id=meet.id,
            ))
        saved += 1

    db.commit()
    return {"saved": saved, "cleared": cleared, "errors": errors, **get_meet_results(meet_id, db)}


@router.post("/{meet_id}/dismiss-results-prompt")
def dismiss_results_prompt(meet_id: int, db: DBSession = Depends(get_db)):
    """Hide this meet's 'add results' prompt from Today without recording anything."""
    meet = _get_or_404(meet_id, db)
    meet.results_prompt_dismissed_at = datetime.now(timezone.utc)
    db.commit()
    return {"dismissed": True, "meet_id": meet.id}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_or_404(meet_id: int, db: DBSession) -> models.Meet:
    meet = db.query(models.Meet).filter(models.Meet.id == meet_id).first()
    if not meet:
        raise HTTPException(status_code=404, detail="Meet not found")
    return meet


def _time_display(seconds: Optional[float]) -> Optional[str]:
    """Render stored seconds the way a coach reads a race time."""
    if seconds is None:
        return None
    minutes = int(seconds // 60)
    remainder = seconds % 60
    return f"{minutes}:{remainder:05.2f}" if minutes else f"{remainder:.2f}"


def _meet_summary(m: models.Meet, db: DBSession) -> dict:
    count = db.query(models.MeetTarget).filter(models.MeetTarget.meet_id == m.id).count()
    return {
        "id": m.id,
        "name": m.name,
        "date": m.date,
        "date_to": m.date_to,
        "location": m.location,
        "course": m.course,
        "level": m.level,
        "swimmer_count": count,
        "timetable_session_count": db.query(models.MeetSession).filter(models.MeetSession.meet_id == m.id).count(),
    }


def _meet_detail(m: models.Meet, db: DBSession) -> dict:
    targets = db.query(models.MeetTarget).filter(models.MeetTarget.meet_id == m.id).all()
    return {
        **_meet_summary(m, db),
        "warm_up_time": m.warm_up_time,
        "notes": m.notes,
        "targets": [_target_out(t, db) for t in targets],
        "timetable": [_session_out(s) for s in m.timetable_sessions],
    }


def _target_out(t: models.MeetTarget, db: DBSession) -> dict:
    swimmer = db.query(models.Swimmer).filter(models.Swimmer.id == t.swimmer_id).first()
    entries = db.query(models.MeetEntry).filter(
        models.MeetEntry.meet_id == t.meet_id,
        models.MeetEntry.swimmer_id == t.swimmer_id,
    ).all()
    return {
        "id": t.id,
        "swimmer_id": t.swimmer_id,
        "swimmer_name": swimmer.name if swimmer else None,
        "squad": swimmer.squad if swimmer else None,
        "events": t.events or [],
        "priority": t.priority,
        "target_times": t.target_times or {},
        "notes": t.notes,
        "scheduled_entries": [{
            "id": entry.id, "event": entry.event_name, "canonical_event": entry.canonical_event,
            "target_time": entry.target_time, "entry_time": entry.entry_time,
            "timetable_linked": bool(entry.meet_event_id),
            "session": entry.meet_event.meet_session.name if entry.meet_event and entry.meet_event.meet_session else None,
            "scheduled_time": entry.meet_event.scheduled_time if entry.meet_event else None,
        } for entry in entries],
    }


def _sync_target_entries(target: models.MeetTarget, db: DBSession, source: str = "manual") -> None:
    desired = {canonicalize_event(event) for event in (target.events or [])}
    existing = db.query(models.MeetEntry).filter(
        models.MeetEntry.meet_id == target.meet_id,
        models.MeetEntry.swimmer_id == target.swimmer_id,
    ).all()
    for entry in existing:
        if entry.canonical_event not in desired:
            db.delete(entry)
    for event in target.events or []:
        upsert_meet_entry(
            target.meet_id, target.swimmer_id, event, db,
            priority=target.priority, target_time=(target.target_times or {}).get(event), source=source,
        )


def _normalise_events(events: list) -> list:
    return [event if isinstance(event, dict) else {"name": str(event)} for event in events]


def _sessions_from_extraction(body: dict, meet: models.Meet) -> list:
    result = []
    for index, (day, events) in enumerate((body.get("by_date") or {}).items()):
        parsed_day = None
        if day != "TBD":
            try:
                parsed_day = datetime.fromisoformat(day).date()
            except ValueError:
                parsed_day = None
        label = parsed_day.strftime("%A %d %B") if parsed_day else f"Competition session {index + 1}"
        result.append({
            "name": label,
            "date": parsed_day.isoformat() if parsed_day else None,
            "order_index": index,
            "events": events,
        })
    if not result and body.get("events"):
        result.append({
            "name": "Competition session",
            "date": meet.date.isoformat() if meet.date else None,
            "order_index": 0,
            "events": body["events"],
        })
    return result


def _session_out(row: models.MeetSession) -> dict:
    return {
        "id": row.id,
        "meet_id": row.meet_id,
        "name": row.name,
        "date": row.date.isoformat() if row.date else None,
        "warm_up_time": row.warm_up_time,
        "start_time": row.start_time,
        "end_time": row.end_time,
        "order_index": row.order_index,
        "events": row.events or [],
        "notes": row.notes,
    }
