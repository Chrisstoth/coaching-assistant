"""
Handles ingestion of swimrankings CSVs and bulk .xlsx session files.
"""
import re
import io
from datetime import datetime, date, time as dt_time, timedelta
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import process as fuzzy_process
from sqlalchemy.orm import Session as DBSession

from backend import models


# ---------------------------------------------------------------------------
# Time helpers
# ---------------------------------------------------------------------------

def parse_time_to_seconds(time_value) -> Optional[float]:
    """Convert Excel times, HH:MM:SS, MM:SS.ss or SS.ss to seconds."""
    if time_value is None or pd.isna(time_value):
        return None
    if isinstance(time_value, dt_time):
        return (
            time_value.hour * 3600
            + time_value.minute * 60
            + time_value.second
            + time_value.microsecond / 1_000_000
        )
    if isinstance(time_value, (timedelta, pd.Timedelta)):
        return float(time_value.total_seconds())
    if isinstance(time_value, (int, float)):
        numeric = float(time_value)
        return numeric * 86400 if 0 < numeric < 1 else numeric

    time_str = str(time_value).strip()
    if ":" in time_str:
        parts = time_str.split(":")
        try:
            if len(parts) == 3:
                return float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2])
            if len(parts) == 2:
                return float(parts[0]) * 60 + float(parts[1])
            return None
        except ValueError:
            return None
    try:
        return float(time_str)
    except ValueError:
        return None


def parse_date_value(value) -> Optional[date]:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    parsed = (
        pd.to_datetime(text, format="%Y-%m-%d", errors="coerce")
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text)
        else pd.to_datetime(text, dayfirst=True, errors="coerce")
    )
    return None if pd.isna(parsed) else parsed.date()


def clean_swimrankings_id(value) -> Optional[str]:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def clean_optional_text(value) -> Optional[str]:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    return text if text and text.lower() != "nan" else None


def infer_distance_from_time(time_seconds: float, stroke: str) -> Optional[int]:
    """
    Rough distance inference from time + stroke.
    Returns most likely distance in metres, or None.
    """
    if time_seconds is None:
        return None
    stroke_lower = (stroke or "").lower()
    # Very rough ranges for each distance/stroke combo
    ranges = {
        50: (20, 80),
        100: (45, 200),
        200: (100, 400),
        400: (230, 600),
        800: (480, 1100),
        1500: (900, 2200),
    }
    # Butterfly and backstroke are slightly slower at 200+
    for dist, (lo, hi) in ranges.items():
        if lo <= time_seconds <= hi:
            return dist
    return None


# ---------------------------------------------------------------------------
# Swimmer matching
# ---------------------------------------------------------------------------

def match_or_create_swimmer(
    name: str,
    swimrankings_id: Optional[str],
    dob: Optional[str],
    gender: Optional[str],
    db: DBSession,
    threshold: int = 85,
) -> models.Swimmer:
    """
    Return existing Swimmer or create a new one.
    Matching priority: swimrankings_id > exact name > fuzzy name.
    """
    # 1. Match by swimrankings ID
    if swimrankings_id:
        existing = db.query(models.Swimmer).filter(
            models.Swimmer.swimrankings_id == str(swimrankings_id)
        ).first()
        if existing:
            return existing

    # 2. Exact name match
    existing = db.query(models.Swimmer).filter(
        models.Swimmer.name == name
    ).first()
    if existing:
        return existing

    # 3. Fuzzy name match
    all_names = [s.name for s in db.query(models.Swimmer.name).all()]
    if all_names:
        result = fuzzy_process.extractOne(name, all_names)
        match, score = (result[0], result[1]) if result else (None, 0)
        if score >= threshold:
            existing = db.query(models.Swimmer).filter(
                models.Swimmer.name == match
            ).first()
            if existing:
                return existing

    # 4. Create new swimmer
    parsed_dob = None
    if dob and not pd.isna(dob):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                parsed_dob = datetime.strptime(str(dob), fmt).date()
                break
            except ValueError:
                continue

    swimmer = models.Swimmer(
        name=name,
        swimrankings_id=str(swimrankings_id) if swimrankings_id and not pd.isna(swimrankings_id) else None,
        dob=parsed_dob,
        gender=str(gender)[0].upper() if gender and not pd.isna(gender) else None,
        active=True,
    )
    db.add(swimmer)
    db.flush()  # get ID without full commit
    return swimmer


# ---------------------------------------------------------------------------
# Swimrankings CSV import
# ---------------------------------------------------------------------------

def import_swimrankings_csv(
    file_content: bytes,
    event_name: str,
    db: DBSession,
) -> dict:
    """
    Parse a swimrankings CSV export and upsert times into the DB.

    Expected columns (subset used):
    Time, WA Pts, Round, Date, Meet, Venue, Club, Level,
    Name, ID, Stroke, Course, Split 1-8, Gender, Age, WA Points/Age, DoB
    """
    df = pd.read_csv(io.BytesIO(file_content))
    df.columns = [c.strip() for c in df.columns]

    results = {"imported": 0, "skipped": 0, "new_swimmers": 0, "errors": []}

    split_cols = [c for c in df.columns if re.match(r"Split \d+", c)]

    for _, row in df.iterrows():
        try:
            name = str(row.get("Name", "")).strip()
            if not name:
                continue

            swimmer = match_or_create_swimmer(
                name=name,
                swimrankings_id=row.get("ID"),
                dob=row.get("DoB"),
                gender=row.get("Gender"),
                db=db,
            )
            is_new = swimmer.id is None
            if is_new:
                results["new_swimmers"] += 1

            time_seconds = parse_time_to_seconds(row.get("Time"))
            if time_seconds is None:
                results["errors"].append(f"Could not parse time for {name}: {row.get('Time')}")
                continue

            stroke = str(row.get("Stroke", "")).strip()
            course = str(row.get("Course", "")).strip().upper()

            # Parse distance from event_name if it starts with a number, else infer from time
            distance = None
            if event_name:
                m = re.match(r'^(\d+)', event_name.strip())
                if m:
                    distance = int(m.group(1))
            if distance is None:
                distance = infer_distance_from_time(time_seconds, stroke)

            # Build canonical event string: always "{base} {course}"
            # Strip any existing course suffix from event_name to avoid duplication
            base = (event_name or f"{distance or '?'} {stroke}").strip()
            for suffix in ('SCM', 'LCM'):
                if base.upper().endswith(suffix):
                    base = base[:-len(suffix)].strip()
            constructed_event = f"{base} {course}".strip() if course else base

            # Parse date
            raw_date = row.get("Date")
            parsed_date = None
            if raw_date and not pd.isna(raw_date):
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        parsed_date = datetime.strptime(str(raw_date), fmt).date()
                        break
                    except ValueError:
                        continue

            # Splits
            splits = []
            for sc in split_cols:
                val = row.get(sc)
                splits.append(parse_time_to_seconds(val) if val and not pd.isna(val) else None)

            # Dedup: same swimmer + event + date + time
            existing_time = db.query(models.SwimTime).filter(
                models.SwimTime.swimmer_id == swimmer.id,
                models.SwimTime.event == constructed_event,
                models.SwimTime.date == parsed_date,
                models.SwimTime.time_seconds == time_seconds,
            ).first()

            if existing_time:
                results["skipped"] += 1
                continue

            wa_pts = None
            raw_wa = row.get("WA Pts")
            if raw_wa and not pd.isna(raw_wa):
                try:
                    wa_pts = float(raw_wa)
                except ValueError:
                    pass

            wa_pts_age = None
            raw_wpa = row.get("WA Points/Age")
            if raw_wpa and not pd.isna(raw_wpa):
                try:
                    wa_pts_age = float(raw_wpa)
                except ValueError:
                    pass

            age_at_swim = None
            raw_age = row.get("Age")
            if raw_age and not pd.isna(raw_age):
                try:
                    age_at_swim = float(raw_age)
                except ValueError:
                    pass

            swim_time = models.SwimTime(
                swimmer_id=swimmer.id,
                event=constructed_event,
                stroke=stroke,
                course=course,
                distance=distance,
                time_seconds=time_seconds,
                wa_points=wa_pts,
                wa_points_age=wa_pts_age,
                date=parsed_date,
                meet=str(row.get("Meet", "")).strip() or None,
                venue=str(row.get("Venue", "")).strip() or None,
                level=str(row.get("Level", "")).strip() or None,
                round=str(row.get("Round", "")).strip() or None,
                splits=splits if any(s is not None for s in splits) else None,
                age_at_swim=age_at_swim,
                source="csv",
            )
            db.add(swim_time)
            results["imported"] += 1

        except Exception as e:
            results["errors"].append(f"Row error ({row.get('Name', '?')}): {str(e)}")

    db.commit()
    return results


# ---------------------------------------------------------------------------
# Combined squad + all-swims workbook
# ---------------------------------------------------------------------------

COMBINED_REQUIRED_COLUMNS = {
    "Time", "Date", "Name", "ID", "Stroke", "Course", "Gender", "DoB",
}


def import_combined_swims_xlsx(
    file_content: bytes,
    squad: str,
    db: DBSession,
    replace_existing: bool = True,
    reconcile_roster: bool = True,
    tracker_content: Optional[bytes] = None,
) -> dict:
    """
    Import one Swim England-style workbook containing the current roster and all swims.

    Each unique ID becomes/updates one swimmer. When replace_existing is true, existing
    race times for swimmers in the workbook are replaced so the workbook is the source
    of truth. Coaching profiles, observations and other swimmer data are preserved.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
    except Exception as exc:
        raise ValueError(f"Could not read workbook: {exc}") from exc

    df.columns = [str(c).strip() for c in df.columns]
    missing = sorted(COMBINED_REQUIRED_COLUMNS - set(df.columns))
    if missing:
        raise ValueError(f"Combined workbook is missing required columns: {', '.join(missing)}")
    if df.empty:
        raise ValueError("Combined workbook contains no swims")

    # Parse and validate all race rows before replacing anything in the database.
    parsed_rows = []
    errors = []
    split_cols = [c for c in df.columns if re.fullmatch(r"Split \d+", c)]
    for row_number, (_, row) in enumerate(df.iterrows(), start=2):
        external_id = clean_swimrankings_id(row.get("ID"))
        name = clean_optional_text(row.get("Name"))
        event_base = clean_optional_text(row.get("Stroke"))
        time_seconds = parse_time_to_seconds(row.get("Time"))
        swim_date = parse_date_value(row.get("Date"))
        raw_course = (clean_optional_text(row.get("Course")) or "").upper()
        course = {"S": "SCM", "SC": "SCM", "SCM": "SCM",
                  "L": "LCM", "LC": "LCM", "LCM": "LCM"}.get(raw_course)

        if not external_id or not name or not event_base or time_seconds is None or not swim_date or not course:
            errors.append(f"Row {row_number}: missing or invalid ID, name, event, time, date, or course")
            continue

        distance_match = re.match(r"^(\d+)\s+(.+)$", event_base)
        distance = int(distance_match.group(1)) if distance_match else infer_distance_from_time(time_seconds, event_base)
        stroke = distance_match.group(2).strip() if distance_match else event_base
        event = f"{event_base} {course}"
        splits = [parse_time_to_seconds(row.get(column)) for column in split_cols]

        def optional_float(column):
            value = row.get(column)
            if value is None or pd.isna(value):
                return None
            try:
                return float(value)
            except (TypeError, ValueError):
                return None

        parsed_rows.append({
            "external_id": external_id,
            "name": name,
            "dob": parse_date_value(row.get("DoB")),
            "gender": clean_optional_text(row.get("Gender")),
            "event": event,
            "stroke": stroke,
            "course": course,
            "distance": distance,
            "time_seconds": time_seconds,
            "wa_points": optional_float("WA Pts"),
            "wa_points_age": optional_float("WA Points/Age"),
            "date": swim_date,
            "meet": clean_optional_text(row.get("Meet")),
            "venue": clean_optional_text(row.get("Venue")),
            "level": clean_optional_text(row.get("Level")),
            "round": clean_optional_text(row.get("Round")),
            "splits": splits if any(value is not None for value in splits) else None,
            "age_at_swim": optional_float("Age"),
        })

    if not parsed_rows:
        raise ValueError("Combined workbook contains no valid swim rows")
    if len(errors) > max(20, int(len(df) * 0.05)):
        raise ValueError(f"Too many invalid rows ({len(errors)} of {len(df)}); no data was changed")

    tracker_by_external_id = {}
    if tracker_content:
        try:
            tracker_df = pd.read_csv(io.BytesIO(tracker_content))
        except Exception as exc:
            raise ValueError(f"Could not read tracker CSV: {exc}") from exc
        tracker_df.columns = [str(c).strip() for c in tracker_df.columns]
        tracker_required = {"ID", "Homeclub", "CS Start Date"}
        tracker_missing = sorted(tracker_required - set(tracker_df.columns))
        if tracker_missing:
            raise ValueError(f"Tracker CSV is missing required columns: {', '.join(tracker_missing)}")
        for _, tracker_row in tracker_df.iterrows():
            external_id = clean_swimrankings_id(tracker_row.get("ID"))
            if external_id:
                tracker_by_external_id[external_id] = {
                    "homeclub": clean_optional_text(tracker_row.get("Homeclub")),
                    "club_start": parse_date_value(tracker_row.get("CS Start Date")),
                }

    squad_label = squad.strip()
    initial_swimmer_ids = {row[0] for row in db.query(models.Swimmer.id).all()}
    reconciliation_candidates = []
    if reconcile_roster and squad_label:
        reconciliation_candidates = (
            db.query(models.Swimmer)
            .filter(models.Swimmer.squad == squad_label, models.Swimmer.active == True)
            .all()
        )
        # One-time migration from the old roster importer, which left squad blank.
        # Only claim unassigned swimmers when the entire active database is unassigned.
        if not reconciliation_candidates:
            active_swimmers = db.query(models.Swimmer).filter(models.Swimmer.active == True).all()
            if active_swimmers and all(not swimmer.squad for swimmer in active_swimmers):
                reconciliation_candidates = active_swimmers
    swimmer_by_external_id = {}
    member_rows = {}
    for row in parsed_rows:
        member_rows.setdefault(row["external_id"], row)

    for external_id, row in member_rows.items():
        swimmer = match_or_create_swimmer(
            name=row["name"],
            swimrankings_id=external_id,
            dob=row["dob"],
            gender=row["gender"],
            db=db,
        )
        # Identity and current squad fields in this workbook are authoritative.
        swimmer.name = row["name"]
        swimmer.swimrankings_id = external_id
        if row["dob"]:
            swimmer.dob = row["dob"]
        if row["gender"]:
            swimmer.gender = row["gender"][0].upper()
        if squad_label:
            swimmer.squad = squad_label
        swimmer.active = True
        if swimmer.status == "inactive":
            swimmer.status = "active"
        tracker = tracker_by_external_id.get(external_id)
        if tracker is not None:
            # Preserve free-form coaching notes while refreshing tracker metadata lines.
            note_lines = [
                line for line in (swimmer.profile_notes or "").splitlines()
                if not line.startswith("Home club swimmer:") and not line.startswith("Squad start:")
            ]
            is_homeclub = (tracker["homeclub"] or "").upper() in {"Y", "YES", "1", "TRUE"}
            note_lines.append(f"Home club swimmer: {'Yes' if is_homeclub else 'No'}")
            if tracker["club_start"]:
                note_lines.append(f"Squad start: {tracker['club_start'].isoformat()}")
            swimmer.profile_notes = "\n".join(line for line in note_lines if line.strip())
        db.flush()
        swimmer_by_external_id[external_id] = swimmer

    swimmer_ids = [swimmer.id for swimmer in swimmer_by_external_id.values()]
    imported_swimmer_ids = set(swimmer_ids)
    marked_inactive = 0
    for swimmer in reconciliation_candidates:
        if swimmer.id not in imported_swimmer_ids:
            swimmer.active = False
            swimmer.status = "inactive"
            if not swimmer.squad:
                swimmer.squad = squad_label
            marked_inactive += 1

    deleted_times = 0
    if replace_existing:
        deleted_times = (
            db.query(models.SwimTime)
            .filter(models.SwimTime.swimmer_id.in_(swimmer_ids))
            .delete(synchronize_session=False)
        )
        known_keys = set()
    else:
        known_keys = {
            (swimmer_id, event, swim_date, time_seconds)
            for swimmer_id, event, swim_date, time_seconds in (
                db.query(
                    models.SwimTime.swimmer_id,
                    models.SwimTime.event,
                    models.SwimTime.date,
                    models.SwimTime.time_seconds,
                )
                .filter(models.SwimTime.swimmer_id.in_(swimmer_ids))
                .all()
            )
        }

    imported = 0
    duplicate_rows = 0
    for row in parsed_rows:
        swimmer = swimmer_by_external_id[row["external_id"]]
        key = (swimmer.id, row["event"], row["date"], row["time_seconds"])
        if key in known_keys:
            duplicate_rows += 1
            continue
        known_keys.add(key)
        db.add(models.SwimTime(
            swimmer_id=swimmer.id,
            event=row["event"],
            stroke=row["stroke"],
            course=row["course"],
            distance=row["distance"],
            time_seconds=row["time_seconds"],
            wa_points=row["wa_points"],
            wa_points_age=row["wa_points_age"],
            date=row["date"],
            meet=row["meet"],
            venue=row["venue"],
            level=row["level"],
            round=row["round"],
            splits=row["splits"],
            age_at_swim=row["age_at_swim"],
            source="combined_xlsx",
        ))
        imported += 1

    db.commit()
    created = sum(swimmer.id not in initial_swimmer_ids for swimmer in swimmer_by_external_id.values())
    return {
        "members_in_file": len(swimmer_by_external_id),
        "swimmers_created": created,
        "swimmers_updated": len(swimmer_by_external_id) - created,
        "swimmers_marked_inactive": marked_inactive,
        "times_imported": imported,
        "times_replaced": deleted_times,
        "duplicate_rows_skipped": duplicate_rows,
        "invalid_rows_skipped": len(errors),
        "errors": errors[:20],
        "squad": squad_label or None,
        "tracker_members_matched": len(set(member_rows) & set(tracker_by_external_id)),
        "tracker_members_missing": len(set(member_rows) - set(tracker_by_external_id)) if tracker_content else 0,
    }


# ---------------------------------------------------------------------------
# Excel session import (single file)
# ---------------------------------------------------------------------------

_MONTHS = {
    name.lower(): number
    for number, name in enumerate(
        ("", "January", "February", "March", "April", "May", "June",
         "July", "August", "September", "October", "November", "December")
    )
}


def _excel_clock(value, *, duration: bool = False) -> Optional[str]:
    seconds = parse_time_to_seconds(value)
    if seconds is None:
        return None
    seconds = int(round(seconds))
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if duration or hours:
        return f"{hours}:{minutes:02d}:{secs:02d}" if hours else f"{minutes}:{secs:02d}"
    return f"{minutes}:{secs:02d}"


def _filename_date(filename: str) -> Optional[date]:
    patterns = (
        (r"(?<!\d)(\d{4}-\d{2}-\d{2})(?!\d)", ("%Y-%m-%d",)),
        (r"(?<!\d)(\d{2}[\-/]\d{2}[\-/]\d{4})(?!\d)", ("%d-%m-%Y", "%d/%m/%Y")),
        # Coach template filenames use YYMMDD: 260824 = 24 August 2026.
        (r"(?<!\d)(\d{6})(?!\d)", ("%y%m%d",)),
    )
    for pattern, formats in patterns:
        match = re.search(pattern, filename)
        if not match:
            continue
        for fmt in formats:
            try:
                return datetime.strptime(match.group(1), fmt).date()
            except ValueError:
                pass
    return None


def _header_date(text: str, year_hint: Optional[int]) -> Optional[date]:
    match = re.search(
        r"\b(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)(?:\s+(\d{4}))?\b",
        text or "",
        re.IGNORECASE,
    )
    if not match:
        return None
    month = _MONTHS.get(match.group(2).lower())
    year = int(match.group(3)) if match.group(3) else year_hint
    if not month or not year:
        return None
    try:
        return date(year, month, int(match.group(1)))
    except ValueError:
        return None


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _template_set_items(ws, start_row: int) -> tuple[list[dict], list[str]]:
    items, raw_lines = [], []
    in_repeat = False
    for row_number in range(start_row, ws.max_row + 1):
        values = [ws.cell(row_number, col).value for col in range(1, 11)]
        reps, _marker, distance, as_word, description, _, sendoff, effort, metres, _duration = values
        if not any(_cell_text(value) for value in values):
            in_repeat = False
            continue

        repeat_match = re.fullmatch(r"\s*(\d+)\s*x\s*", _cell_text(reps), re.IGNORECASE)
        if repeat_match and not any(_cell_text(v) for v in values[1:9]):
            repeat = int(repeat_match.group(1))
            items.append({"type": "repeat", "repetitions": repeat})
            raw_lines.append(f"{repeat}x:")
            in_repeat = True
            continue

        numeric_reps = pd.to_numeric(reps, errors="coerce")
        numeric_distance = pd.to_numeric(distance, errors="coerce")
        if not pd.isna(numeric_reps) and not pd.isna(numeric_distance):
            total_value = pd.to_numeric(metres, errors="coerce")
            item = {
                "type": "set",
                "repetitions": int(numeric_reps),
                "distance": int(numeric_distance),
                "description": _cell_text(description),
                "sendoff": _excel_clock(sendoff),
                "effort": _cell_text(effort) or None,
                "total_metres": None if pd.isna(total_value) else int(total_value),
            }
            items.append(item)
            line = f"{item['repetitions']} x {item['distance']}"
            if _cell_text(as_word):
                line += f" {_cell_text(as_word)}"
            if item["description"]:
                line += f" {item['description']}"
            if item["sendoff"]:
                line += f" @ {item['sendoff']}"
            if item["effort"]:
                line += f" | effort {item['effort']}/20"
            if item["total_metres"] is not None:
                line += f" | {item['total_metres']}m"
            raw_lines.append(("  " if in_repeat else "") + line)
            continue

        note = _cell_text(description) or _cell_text(reps)
        if note:
            items.append({"type": "note", "text": note})
            raw_lines.append(("  " if in_repeat else "") + note)
    return items, raw_lines


def extract_session_xlsx(file_content: bytes, filename: str) -> dict:
    """Extract a reviewable session draft without writing to the database."""
    try:
        workbook = load_workbook(io.BytesIO(file_content), data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError("The workbook could not be opened as an .xlsx file") from exc

    ws = workbook[workbook.sheetnames[0]]
    warnings = []
    title = _cell_text(ws.cell(1, 1).value) or filename.rsplit(".", 1)[0]
    filename_date = _filename_date(filename)
    sheet_date = _header_date(title, filename_date.year if filename_date else date.today().year)
    inferred_date = sheet_date or filename_date
    if sheet_date and filename_date and sheet_date != filename_date:
        warnings.append(
            f"The sheet heading says {sheet_date.isoformat()}, but the filename implies "
            f"{filename_date.isoformat()}; the sheet heading was used."
        )
    if not inferred_date:
        raise ValueError(
            "No session date was found. Use a YYMMDD filename (for example 260824) "
            "or put a date in the sheet heading."
        )

    named_day = re.match(r"\s*(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\b", title, re.I)
    if named_day and named_day.group(1).lower() != inferred_date.strftime("%A").lower():
        warnings.append(
            f"The heading says {named_day.group(1)}, but {inferred_date.isoformat()} is "
            f"{inferred_date.strftime('%A')}."
        )

    start_time = end_time = venue = coach = week_key = None
    aims, set_header_row = [], None
    planned_metres = planned_duration = None
    for row_number in range(1, min(ws.max_row, 20) + 1):
        row_values = [ws.cell(row_number, col).value for col in range(1, 11)]
        row_texts = [_cell_text(value) for value in row_values]
        joined = " | ".join(value for value in row_texts if value)
        time_match = re.search(r"\b(\d{1,2}:\d{2})\s*[-–—]\s*(\d{1,2}:\d{2})\b", joined)
        if time_match:
            start_time, end_time = time_match.groups()
        first = row_texts[0].rstrip(":").strip().lower()
        if row_number == 2 and row_texts[0]:
            venue = row_texts[0]
        elif row_number == 3 and row_texts[0] and ":" not in row_texts[0]:
            coach = row_texts[0]
        if first == "week key":
            week_key = next((value for value in row_texts[1:] if value), None)
        elif re.fullmatch(r"aim\s*\d+", first):
            aim = next((value for value in row_texts[1:] if value), None)
            if aim:
                aims.append(aim)
        if any(value.lower().rstrip(":") in {"warm up", "warm-up", "session", "main set"} for value in row_texts):
            set_header_row = row_number
            numeric_total = pd.to_numeric(ws.cell(row_number, 9).value, errors="coerce")
            if not pd.isna(numeric_total):
                planned_metres = int(numeric_total)
            planned_duration = _excel_clock(ws.cell(row_number, 10).value, duration=True)

    if set_header_row is None:
        raise ValueError("The session set table was not found (expected a Warm Up or Session heading).")
    items, raw_lines = _template_set_items(ws, set_header_row + 1)
    if not any(item["type"] == "set" for item in items):
        raise ValueError("No recognisable set rows were found in the workbook.")

    calculated_metres = sum(item.get("total_metres") or 0 for item in items if item["type"] == "set")
    if planned_metres is not None and calculated_metres and planned_metres != calculated_metres:
        warnings.append(
            f"The workbook total is {planned_metres}m, but the extracted row totals add up to {calculated_metres}m."
        )
    elif planned_metres is None and calculated_metres:
        planned_metres = calculated_metres
    if not start_time:
        warnings.append("No start/end time was found, so the scheduled session cannot be matched automatically.")

    metadata_notes = [f"Imported from: {filename}"]
    metadata_notes += [f"Venue: {venue}"] if venue else []
    metadata_notes += [f"Coach: {coach}"] if coach else []
    metadata_notes += [f"Week key: {week_key}"] if week_key else []
    metadata_notes += [f"Planned total: {planned_metres}m"] if planned_metres is not None else []
    metadata_notes += [f"Planned duration: {planned_duration}"] if planned_duration else []
    draft = {
        "date": inferred_date.isoformat(),
        "start_time": start_time,
        "end_time": end_time,
        "title": title,
        "coach_intent": "\n".join(f"Aim {index}: {aim}" for index, aim in enumerate(aims, 1)) or None,
        "coach_notes": "\n".join(metadata_notes),
        "groups": {
            "1": {
                "description": " · ".join(aims) if aims else "Imported session",
                "sets": "\n".join(raw_lines),
                "items": items,
                "total_metres": planned_metres,
            }
        },
        "source": "excel",
    }
    return {
        "draft": draft,
        "metadata": {
            "filename": filename, "sheet": ws.title, "venue": venue, "coach": coach,
            "week_key": week_key, "aims": aims, "planned_metres": planned_metres,
            "calculated_metres": calculated_metres, "planned_duration": planned_duration,
        },
        "warnings": warnings,
    }


def save_session_xlsx_draft(
    draft: dict,
    db: DBSession,
    target_session_id: Optional[int] = None,
    *,
    mark_historical_complete: bool = False,
) -> models.Session:
    """Create or update a session from a reviewed Excel draft."""
    session_date = date.fromisoformat(str(draft.get("date")))
    pool_slot_id = draft.get("pool_slot_id")
    slot = None
    if pool_slot_id:
        slot = db.query(models.PoolSlot).filter(models.PoolSlot.id == int(pool_slot_id)).first()
        if not slot:
            raise ValueError("The matched timetable slot no longer exists.")
    session = None
    if target_session_id:
        session = db.query(models.Session).filter(models.Session.id == int(target_session_id)).first()
        if not session:
            raise ValueError("The selected session no longer exists.")
    elif slot:
        session = db.query(models.Session).filter(
            models.Session.date == session_date, models.Session.pool_slot_id == slot.id,
        ).first()
    if session and session.status == "cancelled":
        raise ValueError("This scheduled session is cancelled and cannot be overwritten by an import.")
    if session is None:
        status = (
            "completed" if mark_historical_complete and session_date < date.today()
            else "active" if session_date <= date.today()
            else "planned"
        )
        session = models.Session(date=session_date, status=status, source="excel")
        db.add(session)
        db.flush()

    groups = draft.get("groups") or {}
    session.date = session_date
    session.start_time = draft.get("start_time") or (slot.time if slot else session.start_time)
    session.end_time = draft.get("end_time") or (slot.end_time if slot else session.end_time)
    session.squad = draft.get("squad") or (slot.squad if slot else session.squad)
    session.title = draft.get("title") or (slot.label if slot else session.title)
    session.coach_intent = draft.get("coach_intent")
    session.coach_notes = draft.get("coach_notes")
    session.energy_system_focus = draft.get("energy_system_focus")
    session.planned_content = groups
    session.pool_slot_id = slot.id if slot else session.pool_slot_id
    session.course = draft.get("course") or (slot.course if slot else session.course)
    session.source = "excel"

    db.query(models.SessionGroup).filter(models.SessionGroup.session_id == session.id).delete()
    for group_number, content in groups.items():
        total_metres = content.get("total_metres")
        db.add(models.SessionGroup(
            session_id=session.id,
            group_number=int(group_number),
            description=content.get("description", ""),
            sets={"raw": content.get("sets", ""), "items": content.get("items") or []},
            volume_breakdown={"session_total": total_metres} if total_metres else None,
        ))
    db.commit()
    db.refresh(session)
    return session

def import_session_xlsx(
    file_content: bytes,
    filename: str,
    db: DBSession,
) -> dict:
    extracted = extract_session_xlsx(file_content, filename)
    session = save_session_xlsx_draft(extracted["draft"], db, mark_historical_complete=True)
    return {"session_id": session.id, "warnings": extracted["warnings"]}


# ---------------------------------------------------------------------------
# Bulk Excel session import
# ---------------------------------------------------------------------------

def bulk_import_sessions(
    files: list[tuple[str, bytes]],
    db: DBSession,
) -> dict:
    """
    Import multiple .xlsx session files.
    files: list of (filename, bytes) tuples
    """
    summary = {"sessions_created": 0, "errors": [], "warnings": [], "session_ids": []}

    for filename, content in files:
        try:
            result = import_session_xlsx(content, filename, db)
            if result["session_id"]:
                summary["sessions_created"] += 1
                summary["session_ids"].append(result["session_id"])
            summary["warnings"].extend([f"[{filename}] {w}" for w in result["warnings"]])
        except Exception as e:
            summary["errors"].append(f"[{filename}] {str(e)}")

    return summary
